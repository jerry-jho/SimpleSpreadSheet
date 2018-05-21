"""
SimpleSpreadSheet

Read Excel SpreadSheet to comprehensive dataset

DataSet:

{
    'sheet_names' : ['sheet1','sheet2','sheet3],
    'data' : {
         'sheet1' : [
            ['A','B',...], # Row1
            ['D','E',...], # Row2
            ...
         ]
         'sheet2' : ...
    }
}

"""

from openpyxl import load_workbook
from openpyxl import Workbook
import xlrd


def ReadWorkBookXls(filename):
    rtn = {
        'sheet_names' : [],
        'data' : {}
    }
    
    wb = xlrd.open_workbook(filename)
    for _sheetname in wb.sheet_names():
        sheet = wb.sheet_by_name(_sheetname)
        d = []
        for nrow in range(sheet.nrows):
            row_data = []
            row = sheet.row_values(nrow)
            if row:
                for cellv in row:
                    row_data.append(cellv)
            d.append(row_data)
        rtn['sheet_names'].append(_sheetname)
        rtn['data'][_sheetname] = d
    return rtn        


def ReadWorkBookXlsX(filename):
    rtn = {
        'sheet_names' : [],
        'data' : {}
    }
    
    wb = load_workbook(filename)
    for _sheetname in wb.sheetnames:
        sheet = wb[_sheetname]
        d = []
        for row in sheet.rows:
            row_data = []
            for cell in row:    
                if cell.value:
                    v = cell.value
                else:
                    v = ''
                row_data.append(v)
            d.append(row_data)
        rtn['sheet_names'].append(_sheetname)
        rtn['data'][_sheetname] = d
    return rtn

def ReadWorkBook(filename):
    if filename.endswith('.xlsx'):
        return ReadWorkBookXlsX(filename)
    elif filename.endswith('.xls'):
        return ReadWorkBookXls(filename)